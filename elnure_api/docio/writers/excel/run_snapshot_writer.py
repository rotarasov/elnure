from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

from docio.adapters import RunSnapshotDictsAdapter
from docio.writers.base import BaseWriter


class RunSnapshptWriter(BaseWriter):
    def __init__(self):
        self.adapter = RunSnapshotDictsAdapter()

    @property
    def styles(self):
        header = NamedStyle(name="header")
        header.font = Font(size=14, bold=True)
        header.alignment = Alignment(horizontal="center")

        cell = NamedStyle(name="cell")
        cell.font = Font(size=14)

        return {
            "header": header,
            "cell": cell,
        }

    def set_style(self, cell: Cell, styles: dict):
        for attr, style in styles.items():
            setattr(cell, attr, style)

    def build(self, run_snapshot):
        sheets = self.adapter.forward(run_snapshot)

        wb = Workbook()
        for style in self.styles.values():
            wb.add_named_style(style)

        for i, sheet in enumerate(sheets):
            ws = wb.create_sheet(title=sheet.name, index=i)

            offset = 1
            for elective_course, elective_groups in sheet.data.items():
                offset = self.build_elective_course(
                    ws, offset, elective_course, elective_groups
                )

                ws.append([])
                offset += 1

        stream = None
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        return stream

    def build_elective_course(
        self,
        ws: Worksheet,
        offset: int,
        elective_course: str,
        elective_groups: dict[str, list],
    ) -> int:
        """Returning new offset after build"""
        ws.merge_cells(f"A{offset}:D{offset}")
        merged_cell = ws[f"A{offset}"]
        merged_cell.value = elective_course
        merged_cell.style = "header"
        offset += 1

        for elective_group_name, students in elective_groups.items():
            ws.merge_cells(f"A{offset}:D{offset}")
            merged_cell = ws[f"A{offset}"]
            merged_cell.value = elective_group_name
            merged_cell.style = "header"
            offset += 1

            for student in students:
                ws.append(student)
                ws.row_dimensions[offset].styles = "cell"
                offset += 1

        return offset
